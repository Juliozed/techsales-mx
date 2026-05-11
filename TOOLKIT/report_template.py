"""
report_template.py
==================
Reusable automated report template.
Copy into any project. Fill in the sections marked with YOUR DATA HERE.
Run it every time you need a fresh report — the output is always formatted.

WHAT TO CHANGE PER PROJECT:
1. DB_CONFIG — your database credentials
2. PULL DATA section — your table names
3. CLEAN section — your column names and types
4. KPIs section — your business metrics
5. SUMMARIES section — your groupby columns
6. REPORT_NAME — your project name
"""

import pandas as pd
import numpy as np
import sqlalchemy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


# ══════════════════════════════════════════════════════════════════════════
# CONFIGURATION — fill these in for your project
# ══════════════════════════════════════════════════════════════════════════

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "database": "YOUR_DATABASE",  # ← change this
    "user": "postgres",
    "password": "YOUR_PASSWORD",  # ← change this
}

REPORT_NAME = "YOUR_REPORT_NAME"  # ← change this e.g. 'SalesDashboard'
FILTER_YEAR = None  # ← set to e.g. 2024 to filter, None for all years
HEADER_COLOR = "1F3864"  # ← dark navy blue — change if needed


# ══════════════════════════════════════════════════════════════════════════
# STEP 1 — CONNECT
# ══════════════════════════════════════════════════════════════════════════

print("Connecting to database...")

engine = sqlalchemy.create_engine(
    f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
    f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
)


# ══════════════════════════════════════════════════════════════════════════
# STEP 2 — PULL DATA
# Fill in your table names and any SQL filters
# ══════════════════════════════════════════════════════════════════════════

print("Pulling data...")

# ── YOUR DATA HERE ────────────────────────────────────────────
df_main = pd.read_sql("SELECT * FROM your_main_table", engine)
# df_ref1 = pd.read_sql('SELECT * FROM your_reference_table_1', engine)
# df_ref2 = pd.read_sql('SELECT * FROM your_reference_table_2', engine)
# ─────────────────────────────────────────────────────────────

print(f"Pulled {len(df_main):,} rows")


# ══════════════════════════════════════════════════════════════════════════
# STEP 3 — CLEAN
# Fix data types, rename columns, handle nulls, merge reference tables
# ══════════════════════════════════════════════════════════════════════════

print("Cleaning data...")

# ── YOUR DATA HERE ────────────────────────────────────────────

# Fix dates
# df_main['your_date_col'] = pd.to_datetime(df_main['your_date_col'])
# df_main['year']  = df_main['your_date_col'].dt.year
# df_main['month'] = df_main['your_date_col'].dt.month

# Fix numeric columns stored as strings
# df_main['your_amount_col'] = pd.to_numeric(
#     df_main['your_amount_col'].astype(str).str.replace('[$,]','',regex=True),
#     errors='coerce'
# )

# Merge reference tables (LEFT JOIN equivalent)
# df = df_main.merge(df_ref1[['id_col', 'name_col']], on='id_col', how='left')

# Filter to specific year if set
df = df_main.copy()
if FILTER_YEAR:
    df = df[df["year"] == FILTER_YEAR]
    print(f"Filtered to {FILTER_YEAR}: {len(df):,} rows")

# ─────────────────────────────────────────────────────────────

print("Data clean")


# ══════════════════════════════════════════════════════════════════════════
# STEP 4 — CALCULATE KPIs
# Replace with your actual business metrics
# ══════════════════════════════════════════════════════════════════════════

print("Calculating KPIs...")

# ── YOUR DATA HERE ────────────────────────────────────────────

# Example KPIs — replace col names with yours
kpi_1 = df["your_amount_col"].sum()  # e.g. total revenue
kpi_2 = len(df)  # e.g. total orders
kpi_3 = df["your_amount_col"].mean()  # e.g. average order value
kpi_4 = (df["your_status_col"] == "Completed").sum() / len(df) * 100  # completion rate
kpi_5 = (df["your_status_col"] == "Cancelled").sum() / len(df) * 100  # cancel rate

summary_data = pd.DataFrame(
    {
        "KPI": [
            "KPI 1 Label",  # ← change labels
            "KPI 2 Label",
            "KPI 3 Label",
            "KPI 4 Label",
            "KPI 5 Label",
        ],
        "Value": [
            f"${kpi_1:,.2f}",
            f"{kpi_2:,}",
            f"${kpi_3:,.2f}",
            f"{kpi_4:.1f}%",
            f"{kpi_5:.1f}%",
        ],
    }
)

# ─────────────────────────────────────────────────────────────

print("KPIs calculated")


# ══════════════════════════════════════════════════════════════════════════
# STEP 5 — BUILD SUMMARIES
# Replace groupby columns with yours
# ══════════════════════════════════════════════════════════════════════════

# ── YOUR DATA HERE ────────────────────────────────────────────

# Example: summary by category (change 'category_col' and 'amount_col')
summary_1 = (
    df.groupby("your_category_col")
    .agg(
        total=("your_amount_col", "sum"),
        count=("your_id_col", "count"),
        average=("your_amount_col", "mean"),
    )
    .round(2)
    .reset_index()
    .sort_values("total", ascending=False)
)

# Add more summaries as needed:
# summary_2 = df.groupby('another_col').agg(...)
# summary_3 = df.groupby(['col_a', 'col_b']).agg(...)

# ─────────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════════════════════
# STEP 6 — EXPORT TO EXCEL
# Sheet names and data sources — add or remove sheets as needed
# ══════════════════════════════════════════════════════════════════════════

print("Building Excel report...")

report_date = datetime.now().strftime("%Y-%m-%d_%H-%M")
filename = f"{REPORT_NAME}_{report_date}.xlsx"

with pd.ExcelWriter(filename, engine="openpyxl") as writer:
    # ── YOUR DATA HERE — add/remove sheets as needed ──────────
    summary_data.to_excel(writer, sheet_name="Executive Summary", index=False)
    summary_1.to_excel(writer, sheet_name="Summary 1", index=False)
    df.to_excel(writer, sheet_name="Raw Data", index=False)
    # ──────────────────────────────────────────────────────────

print(f"Report saved: {filename}")


# ══════════════════════════════════════════════════════════════════════════
# STEP 7 — FORMAT (runs automatically — no changes needed here)
# ══════════════════════════════════════════════════════════════════════════

print("Formatting report...")

wb = load_workbook(filename)

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
header_align = Alignment(horizontal="center", vertical="center")
data_font = Font(name="Arial", size=10)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Header row formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    # Data row formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 45
        )

    # Freeze header row
    ws.freeze_panes = "A2"

wb.save(filename)
print(f"Formatted report ready: {filename}")
print("Done!")
