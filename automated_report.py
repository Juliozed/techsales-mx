"""
automated_report.py
===================
TechSales MX — Weekly Sales Report
Run this script every Monday morning for a fresh report.
"""

import pandas as pd
import sqlalchemy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── 1. CONNECT & PULL DATA ────────────────────────────────────
print("Connecting to database...")

engine = sqlalchemy.create_engine(
    "postgresql://postgres:gg@localhost:5432/techsales_mx"
)

df_sales = pd.read_sql("SELECT * FROM sales_orders", engine)
df_reps = pd.read_sql("SELECT * FROM reps", engine)
df_products = pd.read_sql("SELECT * FROM products", engine)

print(f"Pulled {len(df_sales)} orders")

# ── 2. CLEAN ──────────────────────────────────────────────────
df_sales["order_date"] = pd.to_datetime(df_sales["order_date"])
df_sales["year"] = df_sales["order_date"].dt.year
df_sales["month"] = df_sales["order_date"].dt.month

# Merge rep names
df = df_sales.merge(
    df_reps[["rep_id", "rep_name", "annual_quota"]], on="rep_id", how="left"
)

# Merge product names
df = df.merge(
    df_products[["product_id", "product_name", "category"]], on="product_id", how="left"
)

print("Data cleaned and merged")

# ── 3. CALCULATE KPIs ─────────────────────────────────────────
print("Calculating KPIs...")

# Company level
total_revenue = df["revenue"].sum()
completed_rev = df[df["status"] == "Completed"]["revenue"].sum()
total_orders = len(df)
completion_rate = (df["status"] == "Completed").sum() / total_orders * 100
cancel_rate = (df["status"] == "Cancelled").sum() / total_orders * 100
avg_order_value = df["revenue"].mean()

# Rep summary
# Rep summary — 2024 only for quota attainment accuracy
df_2024 = df[df["year"] == 2024]

rep_summary = (
    df_2024.groupby("rep_name")
    .agg(
        total_revenue=("revenue", "sum"),
        total_orders=("order_id", "count"),
        completed_rev=(
            "revenue",
            lambda x: x[df_2024.loc[x.index, "status"] == "Completed"].sum(),
        ),
        cancel_rate=("status", lambda x: (x == "Cancelled").sum() / len(x) * 100),
        avg_order_value=("revenue", "mean"),
    )
    .round(2)
    .reset_index()
)

# Merge quota directly instead of aligning by index
quota_map = df_reps.set_index("rep_name")["annual_quota"].to_dict()
rep_summary["quota_attainment"] = rep_summary.apply(
    lambda row: round(
        row["completed_rev"] / quota_map.get(row["rep_name"], 1) * 100, 1
    ),
    axis=1,
)


rep_summary = rep_summary.sort_values("total_revenue", ascending=False)

# Product summary
product_summary = (
    df.groupby(["product_name", "category"])
    .agg(
        total_revenue=("revenue", "sum"),
        total_orders=("order_id", "count"),
        avg_price=("revenue", "mean"),
    )
    .round(2)
    .reset_index()
    .sort_values("total_revenue", ascending=False)
)

# Monthly trend
monthly = (
    df.groupby(["year", "month"])
    .agg(
        revenue=("revenue", "sum"),
        orders=("order_id", "count"),
    )
    .reset_index()
    .sort_values(["year", "month"])
)
monthly["period"] = (
    monthly["year"].astype(str) + "-" + monthly["month"].astype(str).str.zfill(2)
)

print("KPIs calculated")

# ── 4. EXPORT TO EXCEL ────────────────────────────────────────
print("Building Excel report...")

report_date = datetime.now().strftime("%Y-%m-%d_%H-%M")
filename = f"TechSales_Report_{report_date}.xlsx"

summary_data = pd.DataFrame(
    {
        "KPI": [
            "Total Revenue (2024)",
            "Completed Revenue (2024)",
            "Total Orders (2024)",
            "Completion Rate (2024)",
            "Cancellation Rate (2024)",
            "Avg Order Value (2024)",
        ],
        "Value": [
            f"${total_revenue:,.2f}",
            f"${completed_rev:,.2f}",
            f"{total_orders:,}",
            f"{completion_rate:.1f}%",
            f"{cancel_rate:.1f}%",
            f"${avg_order_value:,.2f}",
        ],
    }
)


with pd.ExcelWriter(filename, engine="openpyxl") as writer:
    # Sheet 1 — Executive Summary
    summary_data.to_excel(writer, sheet_name="Executive Summary", index=False)

    # Sheet 2 — Rep Performance
    rep_summary.to_excel(writer, sheet_name="Rep Performance", index=False)

    # Sheet 3 — Product Performance
    product_summary.to_excel(writer, sheet_name="Product Performance", index=False)

    # Sheet 4 — Monthly Trend
    monthly[["period", "revenue", "orders"]].to_excel(
        writer, sheet_name="Monthly Trend", index=False
    )

    # Sheet 5 — Raw Data
    df.to_excel(writer, sheet_name="Raw Data", index=False)

print(f"Report saved: {filename}")

# for a professional look and to ensure the file is properly closed before formatting
# ── 5. FORMAT THE EXCEL REPORT ────────────────────────────────
print("Formatting report...")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = load_workbook(filename)

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F3864")
header_align = Alignment(horizontal="center", vertical="center")
data_font = Font(name="Arial", size=10)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

ws_exec = wb["Executive Summary"]
green_fill = PatternFill("solid", fgColor="E2EFDA")
for row in ws_exec.iter_rows(min_row=2):
    for cell in row:
        cell.fill = green_fill

# ── 6.EMBED CHART INTO EXCEL ────────────────────────────────────
import matplotlib.pyplot as plt
import io
from openpyxl.drawing.image import Image

# Build the chart
fig, ax = plt.subplots(figsize=(10, 5))

rep_chart = rep_summary.sort_values("total_revenue", ascending=False)
bars = ax.bar(
    rep_chart["rep_name"],
    rep_chart["total_revenue"],
    color="#2E75B6",
    edgecolor="white",
    width=0.6,
)

# Value labels on bars
for bar in bars:
    height = bar.get_height()
    ax.text(
        bar.get_x() + bar.get_width() / 2,
        height * 1.01,
        f"${height / 1_000_000:.1f}M",
        ha="center",
        va="bottom",
        fontsize=10,
        fontweight="bold",
        color="#1F3864",
    )

ax.set_title(
    "Revenue by Sales Rep — 2024",
    fontsize=13,
    fontweight="bold",
    color="#1F3864",
    pad=15,
)
ax.set_xlabel("Sales Rep", fontsize=11)
ax.set_ylabel("Revenue (MXN)", fontsize=11)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)
ax.grid(axis="y", alpha=0.3)
plt.tight_layout()

# Save chart to memory buffer — no temp file needed
buffer = io.BytesIO()
plt.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
buffer.seek(0)
plt.close()

# Insert chart into Executive Summary sheet
ws_exec = wb["Executive Summary"]
img = Image(buffer)
img.anchor = "D2"  # ← cell where chart appears — change if needed
ws_exec.add_image(img)

print("Chart embedded in Executive Summary")


##-----------------------------last part_____________________________________________________
wb.save(filename)
print(f"Formatted report saved: {filename}")
print("Done!")
